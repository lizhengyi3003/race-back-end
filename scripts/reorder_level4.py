# -*- coding: utf-8 -*-
"""把第4级「具体营业类型」指标按所属小类插入到对应小类之后，
恢复「大类 → 中类 → 小类 → 具体营业类型」的树形层次结构。
输出到 _reordered 临时文件，验证后替换原文件。"""
import io
import shutil
import openpyxl
from collections import OrderedDict

XLSX = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.xlsx"
MD = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.md"
XLSX_OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_reordered.xlsx"
MD_OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_reordered.md"

COMMENT_MARK = "第4级"


def small_code(cat: str) -> str:
    return cat.split("_")[0].split(" ")[0].strip()


def reorder(original_rows, level4_rows):
    """original_rows: 大类/中类/小类行(有序)。level4_rows: 具体营业类型行(有序)。"""
    last_pos = {}
    for i, r in enumerate(original_rows):
        if r[0] == "小类":
            last_pos[small_code(r[1])] = i
    groups = OrderedDict()
    for r in level4_rows:
        groups.setdefault(small_code(r[1]), []).append(r)
    insert = {}
    for code, grp in groups.items():
        pos = last_pos.get(code)
        if pos is not None:
            insert.setdefault(pos, []).extend(grp)
        else:
            print(f"[警告] 找不到所属小类: {code}, {len(grp)} 条被丢弃")
    out = []
    for i, r in enumerate(original_rows):
        out.append(r)
        if i in insert:
            out.extend(insert[i])
    return out


# ==================== XLSX ====================
print("== 处理 XLSX ==")
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
hdr = [c.value for c in ws[1]]
rows = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]

original = [r for r in rows if r[0] != "具体营业类型"]
level4 = [r for r in rows if r[0] == "具体营业类型"]
print(f"原有行 {len(original)}, 第4级 {len(level4)}")
new_rows = reorder(original, level4)
print(f"重组后总行 {len(new_rows)} (应为 {len(original) + len(level4)})")

# 重建：保留表头样式与列宽，清空数据行后按新顺序写入
ws.delete_rows(2, ws.max_row)
for r in new_rows:
    ws.append(r)
wb.save(XLSX_OUT)
print(f"已保存 {XLSX_OUT}")

# ==================== MD ====================
print("== 处理 MD ==")
with io.open(MD, encoding="utf-8") as f:
    lines = f.readlines()

# 找注释行
comment_idx = None
for i, ln in enumerate(lines):
    if COMMENT_MARK in ln and ln.lstrip().startswith("<!--"):
        comment_idx = i
        break
if comment_idx is None:
    print("[错误] 未找到第4级注释行，中止 MD 处理")
else:
    # 数据行解析
    def parse_table(ln):
        s = ln.strip()
        if not (s.startswith("|") and s.endswith("|")):
            return None
        if s.startswith("| ---") or s.startswith("|---"):
            return None
        parts = [p.strip() for p in s.strip("|").split("|")]
        if len(parts) < 14 or parts[0] in ("层级", ""):
            return None
        return parts[:14]

    # 原表格区 = 第一个数据行 到 comment 之前
    first_data = None
    for i, ln in enumerate(lines):
        if parse_table(ln) and parse_table(ln)[0] in ("大类", "中类", "小类"):
            first_data = i
            break
    head = lines[:first_data]

    body = []
    for ln in lines[first_data:comment_idx]:
        p = parse_table(ln)
        if p and p[0] in ("大类", "中类", "小类"):
            body.append(p)

    l4 = []
    for ln in lines[comment_idx + 1:]:
        p = parse_table(ln)
        if p and p[0] == "具体营业类型":
            l4.append(p)

    print(f"头部行 {len(head)}, 原有表格行 {len(body)}, 第4级 {len(l4)}")
    new_body = reorder(body, l4)
    print(f"重组后表格行 {len(new_body)} (应为 {len(body) + len(l4)})")

    def to_md_line(r):
        return "| " + " | ".join(str(x) for x in r) + " |\n"

    with io.open(MD_OUT, "w", encoding="utf-8", newline="\n") as f:
        for ln in head:
            f.write(ln)
        for r in new_body:
            f.write(to_md_line(r))
        f.write("\n<!-- ============ 第4级「具体营业类型」指标（共 2223 条，已按所属小类归入上述树形结构，自动生成 2026-08-07） ============ -->\n")
    print(f"已保存 {MD_OUT}")

print("完成")
