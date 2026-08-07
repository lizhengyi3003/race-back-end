# -*- coding: utf-8 -*-
"""根据检查报告采纳意见：
1. 问题2：第4级「播种面积/存栏量/养殖面积」核心规模指标权重 ★★ -> ★★★
2. 问题1：为 8 个明确的具体营业类型补充特色指标（每类 2 条，共 16 条）
同步修改 xlsx（保留颜色样式）与 md，输出到 _v2 临时文件。"""
import io
from copy import copy
import openpyxl

XLSX = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.xlsx"
MD = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.md"
XLSX_OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_v2.xlsx"
MD_OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_v2.md"

# 新增特色指标（14列）：层级, 所属类别, 指标名称, 指标类型, 单位, 取值说明,
# 数据来源, 是否特色指标, 风险含义, 建议权重, 适用区域, 是否一票否决, 采集周期, 评分规则
NEW_INDICATORS = {
    "0113_0132": [
        ["具体营业类型", "0113_0132 麻类种植", "麻类纤维等级合格率", "数值", "%",
         "抽检纤维等级达标批次占比", "质检记录", "否", "原料品质与销路", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
        ["具体营业类型", "0113_0132 麻类种植", "亚麻订单种植面积", "数值", "亩",
         "亚麻订单/基地合同种植面积", "购销合同", "是", "黑龙江亚麻原料保障", "★★", "黑龙江", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
    "0113_0133": [
        ["具体营业类型", "0113_0133 糖料种植", "甜菜含糖率", "数值", "%",
         "当年平均含糖率（东北近年参考约16-18%）", "制糖企业质检", "是", "品质与收购价", "★★", "黑龙江+蒙东", "否", "年报",
         "数值越高得分越高（分段计分）"],
        ["具体营业类型", "0113_0133 糖料种植", "糖料订单收购比例", "数值", "%",
         "订单收购量占比", "收购合同", "否", "销路保障", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
    "0114_0143": [
        ["具体营业类型", "0114_0143 花卉种植", "设施栽培面积占比", "数值", "%",
         "温室/大棚花卉面积占比", "现场核查/设施台账", "是", "寒地花卉设施投入", "★★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
        ["具体营业类型", "0114_0143 花卉种植", "花卉商品率", "数值", "%",
         "可售商品花占比", "销售记录", "否", "品质与销路", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
    "0116_0162": [
        ["具体营业类型", "0116_0162 含油果种植", "含油果含油率", "数值", "%",
         "果实含油率抽检均值", "质检记录", "否", "原料品质", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
        ["具体营业类型", "0116_0162 含油果种植", "含油果加工转化率", "数值", "%",
         "就地加工/销售占比", "生产台账", "否", "增值与销路", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
    "0118_0182": [
        ["具体营业类型", "0118_0182 天然草原割草", "草场载畜量", "数值", "羊单位/亩",
         "草场实际载畜强度", "牧业台账", "否", "草场退化风险", "★", "东北全境", "否", "年报",
         "数值越高风险越高（分段扣分）"],
        ["具体营业类型", "0118_0182 天然草原割草", "天然草场退化程度", "枚举", "—",
         "无/轻/中/重", "草原监测部门", "否", "资源可持续风险", "★", "东北全境", "否", "年报",
         "按档位映射得分（越好分值越高）"],
    ],
    "0131_0312": [
        ["具体营业类型", "0131_0312 马的饲养", "马匹用途结构", "枚举", "—",
         "肉用/役用/运动/奶用为主", "养殖档案", "否", "销路与价值分化", "★", "东北全境", "否", "年报",
         "按档位映射得分（越好分值越高）"],
        ["具体营业类型", "0131_0312 马的饲养", "马匹良种覆盖率", "数值", "%",
         "良种/纯种马占比", "良种档案", "否", "繁殖与价值", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
    "0131_0315": [
        ["具体营业类型", "0131_0315 骆驼饲养", "骆驼产品结构", "枚举", "—",
         "乳用/肉用/毛绒用为主", "养殖档案", "否", "产品价值与销路", "★", "东北全境", "否", "年报",
         "按档位映射得分（越好分值越高）"],
        ["具体营业类型", "0131_0315 骆驼饲养", "驼绒/驼奶产出率", "数值", "%",
         "单位产出水平", "养殖档案", "否", "产出效率", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
    "0132_0329": [
        ["具体营业类型", "0132_0329 其他家禽饲养", "特色禽类品种", "枚举", "—",
         "鸵鸟/孔雀/鹌鹑/肉鸽等", "养殖档案/许可", "否", "品种价值与销路", "★", "东北全境", "否", "年报",
         "按档位映射得分（越好分值越高）"],
        ["具体营业类型", "0132_0329 其他家禽饲养", "特色禽类产值占比", "数值", "%",
         "特色禽类收入占比", "销售台账", "否", "收入结构", "★", "东北全境", "否", "年报",
         "数值越高得分越高（分段计分）"],
    ],
}


def is_core_scale(name: str) -> bool:
    return name.endswith("播种面积") or name.endswith("存栏量") or name.endswith("养殖面积")


# ==================== XLSX ====================
print("== XLSX ==")
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
rows = [list(r) for r in ws.iter_rows(min_row=2)]

# 权重提升
weight_up = 0
for r in rows:
    if r[0].value == "具体营业类型":
        name = str(r[2].value or "")
        if is_core_scale(name) and r[9].value == "★★":
            r[9].value = "★★★"
            weight_up += 1
print(f"权重提升 {weight_up} 条")

# 样式模板（第4级某行）
template = None
for r in rows:
    if r[0].value == "具体营业类型":
        template = r
        break
template_styles = [copy(c._style) for c in template]

# 找每个具体营业类型最后出现位置
last_idx = {}
for i, r in enumerate(rows):
    if r[0].value == "具体营业类型":
        last_idx[str(r[1].value).split(" ")[0]] = i

# 重建（含插入）
final = []
inserted = 0
for i, r in enumerate(rows):
    final.append(r)
    code = str(r[1].value).split(" ")[0] if r[0].value == "具体营业类型" else None
    if code in NEW_INDICATORS and last_idx.get(code) == i:
        for vals in NEW_INDICATORS[code]:
            final.append(list(zip(vals, template_styles)))
            inserted += 1
print(f"插入特色指标 {inserted} 条 (应 16)")

# 写回（保留样式）
ws.delete_rows(2, ws.max_row)
row_idx = 2
for r in final:
    if isinstance(r[0], tuple):  # 新插入行 (value, style)
        for c, (val, sty) in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell._style = copy(sty)
    else:  # 原有行 cell 对象
        for c, cell in enumerate(r, start=1):
            new_cell = ws.cell(row=row_idx, column=c, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
    row_idx += 1
wb.save(XLSX_OUT)
print(f"已保存 {XLSX_OUT}, 总数据行 {row_idx - 2}")

# ==================== MD ====================
print("== MD ==")
with io.open(MD, encoding="utf-8") as f:
    lines = f.readlines()

def parse_table(ln):
    s = ln.strip()
    if not (s.startswith("|") and s.endswith("|")):
        return None
    if s.startswith("|---"):
        return None
    p = [x.strip() for x in s.strip("|").split("|")]
    if len(p) < 14 or p[0] in ("层级", ""):
        return None
    return p[:14]

def to_md_line(r):
    return "| " + " | ".join(str(x) for x in r) + " |\n"

# 找注释行
comment_idx = None
for i, ln in enumerate(lines):
    if "第4级" in ln and ln.lstrip().startswith("<!--"):
        comment_idx = i
        break
if comment_idx is None:
    print("[错误] md 未找到第4级注释")
else:
    # 处理数据区：first_data ~ comment 之前
    first_data = None
    for i, ln in enumerate(lines):
        p = parse_table(ln)
        if p and p[0] in ("大类", "中类", "小类"):
            first_data = i
            break
    head = lines[:first_data]

    # 收集数据行（带 md 原始行）
    data = []
    for i in range(first_data, comment_idx):
        p = parse_table(lines[i])
        if p:
            data.append(p)
    # 注意：head 到 comment 之间应只有 大类/中类/小类（第4级已归位，在中间）
    # 重新收集：所有表格数据行都要保留顺序。这里 data 含大类/中类/小类/具体营业类型
    # 但 first_data~comment 只到注释，注释前的数据行 = 全部原有 + 归位的第4级
    # 为简单：解析整个 first_data..end 的表格行
    all_data = []
    for i in range(first_data, len(lines)):
        p = parse_table(lines[i])
        if p and p[0] in ("大类", "中类", "小类", "具体营业类型"):
            all_data.append(p)
        if i == comment_idx:
            pass

    # 权重提升 + 找最后位置
    for r in all_data:
        if r[0] == "具体营业类型" and is_core_scale(r[2]) and r[9] == "★★":
            r[9] = "★★★"
    last_idx_md = {}
    for i, r in enumerate(all_data):
        if r[0] == "具体营业类型":
            last_idx_md[str(r[1]).split(" ")[0]] = i

    final_md = []
    ins_md = 0
    for i, r in enumerate(all_data):
        final_md.append(r)
        code = str(r[1]).split(" ")[0] if r[0] == "具体营业类型" else None
        if code in NEW_INDICATORS and last_idx_md.get(code) == i:
            for vals in NEW_INDICATORS[code]:
                final_md.append(list(vals))
                ins_md += 1
    print(f"MD 插入 {ins_md} 条, 表格数据行 {len(final_md)}")

    new_total = len([r for r in final_md if r[0] == "具体营业类型"])
    print(f"MD 第4级总数 {new_total}")

    with io.open(MD_OUT, "w", encoding="utf-8", newline="\n") as f:
        for ln in head:
            f.write(ln)
        for r in final_md:
            f.write(to_md_line(r))
        f.write(f"\n<!-- ============ 第4级「具体营业类型」指标（共 {new_total} 条，已按所属小类归入上述树形结构，自动生成 2026-08-07） ============ -->\n")
    print(f"已保存 {MD_OUT}")

print("完成")
