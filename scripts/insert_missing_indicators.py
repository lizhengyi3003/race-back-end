# -*- coding: utf-8 -*-
"""补充检查报告建议的重要指标（6 条）：
- 基本项: 农产品商标注册情况
- 大类01: 土地流转面积 / 土地流转剩余年限 / 农业社会化服务覆盖率
- 小类0111: 粮食总产量 / 粮食烘干仓储能力
输出到 _v4 临时文件，保留各层级颜色样式。"""
import io
from copy import copy
import openpyxl

XLSX = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.xlsx"
MD = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.md"
XLSX_OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_v4.xlsx"
MD_OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_v4.md"

# 新指标（14列）
B_ITEMS = [
    ["基本项", "—（通用）", "农产品商标注册情况", "枚举", "—",
     "已注册商标/申报中/无", "商标注册证/受理通知", "否", "品牌化经营与溢价",
     "★★★", "东北全境", "否", "年报", "按档位映射得分（越好分值越高）"],
]
D01_ITEMS = [
    ["大类", "01 农林牧渔业", "土地流转面积", "数值", "亩",
     "流转/承包土地面积（不含自有）", "土地流转合同", "否", "经营扩张与稳定性",
     "★★★", "东北全境", "否", "年报", "数值越高得分越高（分段计分）"],
    ["大类", "01 农林牧渔业", "土地流转剩余年限", "数值", "年",
     "主要流转合同剩余年限（加权平均）", "土地流转合同", "否", "经营稳定性与投入沉淀",
     "★★★", "东北全境", "否", "年报", "数值越高得分越高（分段计分）"],
    ["大类", "01 农林牧渔业", "农业社会化服务覆盖率", "数值", "%",
     "接受农机/植保/托管等社会化服务面积占比", "服务合同/作业记录", "否", "小农户与现代农业衔接",
     "★★★", "东北全境", "否", "年报", "数值越高得分越高（分段计分）"],
]
S0111_ITEMS = [
    ["小类", "0111 谷物种植", "粮食总产量", "数值", "吨",
     "近12个月粮食（谷物+豆类+薯类折粮）总产量", "购销记录/现场核查", "否", "生产规模与稳定性",
     "★★★", "东北全境", "否", "年报", "数值越高得分越高（分段计分）"],
    ["小类", "0111 谷物种植", "粮食烘干仓储能力", "数值", "吨",
     "自有+租赁烘干塔/粮仓有效容量（秋收后粮食安全存储关键）", "现场核查/设备台账", "否", "产后损耗与售粮节奏",
     "★★★", "东北全境", "否", "年报", "数值越高得分越高（分段计分）"],
]


def find_template(rows, lv_filter, cat_prefix=None):
    for r in rows:
        lv = r[0].value
        cat = str(r[1].value or "")
        if lv == lv_filter and (cat_prefix is None or cat.startswith(cat_prefix)):
            return r
    return None


# ==================== XLSX ====================
print("== XLSX ==")
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
rows = [list(r) for r in ws.iter_rows(min_row=2)]

tpl_basic = find_template(rows, "基本项")
tpl_big = find_template(rows, "大类", "01 ")
tpl_small = find_template(rows, "小类", "0111 ")

# 找插入位置（各层级最后一行）
last_idx = {}
for i, r in enumerate(rows):
    lv = r[0].value
    cat = str(r[1].value or "")
    if lv == "基本项":
        last_idx["基本项"] = i
    elif lv == "大类" and cat.startswith("01 "):
        last_idx["大类01"] = i
    elif lv == "小类" and cat.startswith("0111 "):
        last_idx["小类0111"] = i
print("插入位置:", last_idx)

insert_map = {
    last_idx["基本项"]: (B_ITEMS, tpl_basic),
    last_idx["大类01"]: (D01_ITEMS, tpl_big),
    last_idx["小类0111"]: (S0111_ITEMS, tpl_small),
}

# 重建
final = []
inserted = 0
for i, r in enumerate(rows):
    final.append(r)
    if i in insert_map:
        items, tpl = insert_map[i]
        tpl_styles = [copy(c._style) for c in tpl]
        for vals in items:
            final.append(list(zip(vals, tpl_styles)))
            inserted += 1
print(f"插入 {inserted} 条 (应 6)")

ws.delete_rows(2, ws.max_row)
row_idx = 2
for r in final:
    if isinstance(r[0], tuple):
        for c, (val, sty) in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell._style = copy(sty)
    else:
        for c, cell in enumerate(r, start=1):
            new_cell = ws.cell(row=row_idx, column=c, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
    row_idx += 1
wb.save(XLSX_OUT)
print(f"已保存 {XLSX_OUT}, 总数据行 {row_idx - 2}")

# ==================== MD ====================
print("== MD ==")
with io.open(MD, encoding="utf-8") as f:
    lines = f.readlines()

def parse_table(ln):
    s = ln.strip()
    if not (s.startswith("|") and s.endswith("|")):
        return None
    if s.startswith("|---"):
        return None
    p = [x.strip() for x in s.strip("|").split("|")]
    if len(p) < 14 or p[0] in ("层级", ""):
        return None
    return p[:14]

def to_md_line(r):
    return "| " + " | ".join(str(x) for x in r) + " |\n"

# 找注释行
comment_idx = None
for i, ln in enumerate(lines):
    if "第4级" in ln and ln.lstrip().startswith("<!--"):
        comment_idx = i
        break

first_data = None
for i, ln in enumerate(lines):
    p = parse_table(ln)
    if p and p[0] in ("大类", "中类", "小类", "基本项"):
        first_data = i
        break
head = lines[:first_data]

# 数据区 = first_data..comment
data_lines = lines[first_data:comment_idx]

last_idx_md = {}
for i, ln in enumerate(data_lines):
    p = parse_table(ln)
    if not p:
        continue
    lv = p[0]; cat = p[1]
    if lv == "基本项":
        last_idx_md["基本项"] = i
    elif lv == "大类" and cat.startswith("01 "):
        last_idx_md["大类01"] = i
    elif lv == "小类" and cat.startswith("0111 "):
        last_idx_md["小类0111"] = i

insert_map_md = {
    last_idx_md["基本项"]: B_ITEMS,
    last_idx_md["大类01"]: D01_ITEMS,
    last_idx_md["小类0111"]: S0111_ITEMS,
}

out = []
for i, ln in enumerate(data_lines):
    out.append(ln)
    if i in insert_map_md:
        for vals in insert_map_md[i]:
            out.append(to_md_line(vals))

with io.open(MD_OUT, "w", encoding="utf-8", newline="\n") as f:
    for ln in head:
        f.write(ln)
    for ln in out:
        f.write(ln)
    for ln in lines[comment_idx:]:
        f.write(ln)
print(f"已保存 {MD_OUT}")
print("完成")
