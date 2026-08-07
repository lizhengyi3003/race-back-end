"""命令行：解析 docs 指标体系 xlsx → 写入 indicator_category + indicator_config。

用法:
    python scripts/import_indicators.py
    python scripts/import_indicators.py --xlsx "path/to/xxx.xlsx"
    python scripts/import_indicators.py --check-only   # 仅解析并打印统计，不写库

数据源: fore-end/docs/农业及相关产业动态指标搜集体系.xlsx（776 行 = 775 数据 + 表头，14 列）。
"""

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from app.db.base import Base  # noqa: E402
from app.db.session import SessionLocal, engine  # noqa: E402
from app.models.indicator import IndicatorCategory, IndicatorConfig  # noqa: E402

DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent.parent
    / "fore-end"
    / "docs"
    / "农业及相关产业动态指标搜集体系.xlsx"
)

HEADER_EXPECTED = [
    "层级", "所属类别", "指标名称", "指标类型", "单位", "取值说明", "数据来源",
    "是否特色指标", "风险含义", "建议权重", "适用区域", "是否一票否决", "采集周期", "评分规则",
]


def parse_star(value: str) -> float:
    """星级 → 数值：★=1, ☆=0.5。如 ★★★★☆ = 4.5。"""
    v = str(value or "").strip()
    if not v:
        return 3.0
    return v.count("★") + v.count("☆") * 0.5


def split_category(value: str) -> tuple[str, str]:
    """所属类别 '0111 谷物种植' → ('0111', '谷物种植')；基本项 → ('BASIC', '—（通用）')。"""
    v = str(value or "").strip()
    if not v or v.startswith("—"):
        return "BASIC", v or "—（通用）"
    parts = v.split(" ", 1)
    code = parts[0].strip()
    name = parts[1].strip() if len(parts) > 1 else code
    return code, name


def derive_parent(code: str, level: str) -> str | None:
    """类别父编码：中类(3位)父=前2位；小类(4位)父=前3位；
    具体营业类型(小类码_行业码)父=小类码；大类无父。"""
    if level == "具体营业类型":
        return code.split("_")[0] if "_" in code else (code[:4] if len(code) >= 4 else None)
    if level == "中类" and len(code) >= 3:
        return code[:2]
    if level == "小类" and len(code) >= 4:
        return code[:3]
    return None


def parse_enum_options(value_range: str) -> list[str]:
    """从取值说明解析枚举档位选项（去掉括号/分号后的说明文字）。"""
    s = str(value_range or "")
    s = re.sub(r"[（(].*?[）)]", "", s)          # 去括号说明
    s = re.split(r"[；;]", s)[0]                  # 去分号说明
    for sep in ("/", "／", "、", "，"):
        if sep in s:
            return [p.strip() for p in s.split(sep) if p.strip()]
    return []


def build_enum_map(options: list[str], name: str) -> dict | None:
    """枚举档位 → 分值映射（scoring_config.map）。
    绝大多数枚举按『从好到坏』书写（稳定/一般/不稳定、无/轻/中/重…），
    少数『认证/地标/商标/林下参』从无到有是『从坏到好』；结构/用途/品种类为中性不配置。"""
    opts = [o.strip() for o in options if o.strip()]
    if len(opts) < 2:
        return None
    # 特例：银行信贷履约记录（无逾期/有逾期/逾期已结清）——有逾期最差，已结清中等
    if opts[0] == "无逾期" or (opts[0] == "无" and len(opts) == 3 and "逾期" in " ".join(opts)):
        return {opts[0]: 100, opts[1]: 30, opts[2]: 60}
    # 中性档位（用途/结构/品种/`xx为主`类）：无好坏之分，不配置（引擎给中分）
    if any(k in name for k in ("用途", "结构", "品种")) or any("为主" in o for o in opts):
        return None
    # 从坏到好（起点=最差）：认证/地标/商标/林下参经营 等
    joined = " ".join(opts)
    if opts[0] == "无" and any(k in joined for k in ("申报中", "省级", "国家级", "规模化", "少量", "已认证")):
        n = len(opts)
        return {opts[i]: round(20 + 80 * i / (n - 1)) for i in range(n)}  # 20,60,100
    # 默认从好到坏：100, 63, 25 / 100, 75, 50, 25
    n = len(opts)
    return {opts[i]: round(100 - 75 * i / (n - 1)) for i in range(n)}


def build_numeric_config(scoring_rule: str) -> dict | None:
    """数值指标方向：规则含『风险越高/越高越差』→ lower_better=true；
    『得分越高/越高越好』→ lower_better=false。显式配置优先于引擎启发。"""
    r = scoring_rule or ""
    if any(k in r for k in ("风险越高", "越高风险", "越高越差", "越高越不利", "越低越好")):
        return {"lower_better": True}
    if any(k in r for k in ("得分越高", "越高越好", "越高分越高")):
        return {"lower_better": False}
    return None


def parse_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(c or "").strip() for c in rows[0]]
    if header != HEADER_EXPECTED:
        print(f"[warn] 表头与预期不完全一致: {header}")
    data = [r for r in rows[1:] if r and any(c is not None and str(c).strip() for c in r)]
    return data


def build(data: list[dict]) -> tuple[list[dict], list[dict]]:
    """返回 (categories, indicators)。"""
    categories: dict[str, dict] = {}
    indicators: list[dict] = []
    basic_seq = 0
    cat_seq: dict[str, int] = {}

    for idx, r in enumerate(data):
        level = str(r[0] or "").strip()
        cat_code, cat_name = split_category(r[1])
        name = str(r[2] or "").strip()
        ind_type = str(r[3] or "数值").strip() or "数值"
        unit = str(r[4] or "").strip()
        value_range = str(r[5] or "").strip()
        data_source = str(r[6] or "").strip()
        is_feature = str(r[7] or "").strip() == "是"
        risk_meaning = str(r[8] or "").strip()
        weight_star = parse_star(r[9])
        region = str(r[10] or "").strip()
        is_veto = str(r[11] or "").strip() == "是"
        cycle = str(r[12] or "").strip()
        scoring_rule = str(r[13] or "").strip()

        # 评分参数：枚举档位 map / 数值方向，显式配置优先于引擎关键词启发
        scoring_config = None
        if ind_type == "枚举":
            _opts = parse_enum_options(value_range)
            _m = build_enum_map(_opts, name) if _opts else None
            if _m:
                scoring_config = {"map": _m}
        elif ind_type == "数值":
            scoring_config = build_numeric_config(scoring_rule)

        # 类别节点（基本项作为伪类别 BASIC）
        key = cat_code
        if cat_code != "BASIC":
            categories.setdefault(key, {
                "code": cat_code,
                "name": cat_name,
                "level": level,
                "parent_code": derive_parent(cat_code, level),
                "display_order": len(categories),
            })

        # 指标编码
        if cat_code == "BASIC":
            basic_seq += 1
            ind_code = f"BASIC_{basic_seq:03d}"
        else:
            cat_seq[cat_code] = cat_seq.get(cat_code, 0) + 1
            ind_code = f"{cat_code}_{cat_seq[cat_code]:02d}"

        indicators.append({
            "code": ind_code,
            "name": name,
            "level": level,
            "category_code": cat_code,
            "category_name": cat_name,
            "indicator_type": ind_type,
            "unit": unit,
            "value_range": value_range,
            "data_source": data_source,
            "is_feature": is_feature,
            "risk_meaning": risk_meaning,
            "weight_star": weight_star,
            "region": region,
            "is_veto": is_veto,
            "cycle": cycle,
            "scoring_rule": scoring_rule,
            "scoring_config": scoring_config,
            "display_order": idx,
        })

    return list(categories.values()), indicators


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--check-only", action="store_true")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[err] 文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"解析: {xlsx_path.name}")
    data = parse_rows(xlsx_path)
    categories, indicators = build(data)

    # 统计
    from collections import Counter

    level_cnt = Counter(i["level"] for i in indicators)
    feature_cnt = sum(1 for i in indicators if i["is_feature"])
    veto_cnt = sum(1 for i in indicators if i["is_veto"])
    print(f"指标总数: {len(indicators)}")
    print(f"  按层级: {dict(level_cnt)}")
    print(f"  特色指标: {feature_cnt}, 一票否决: {veto_cnt}")
    print(f"类别节点: {len(categories)} (大类 {sum(1 for c in categories if c['level']=='大类')} / "
          f"中类 {sum(1 for c in categories if c['level']=='中类')} / 小类 {sum(1 for c in categories if c['level']=='小类')} / "
          f"具体营业类型 {sum(1 for c in categories if c['level']=='具体营业类型')})")

    # 校验唯一性
    codes = [i["code"] for i in indicators]
    if len(set(codes)) != len(codes):
        dup = [c for c in set(codes) if codes.count(c) > 1]
        print(f"[err] 指标编码重复: {dup}")
        sys.exit(1)

    if args.check_only:
        print("[check-only] 未写库。")
        return

    # 写库
    Base.metadata.create_all(engine)
    with SessionLocal() as db:
        db.query(IndicatorConfig).delete()
        db.query(IndicatorCategory).delete()
        db.add_all([IndicatorCategory(**c) for c in categories])
        db.add_all([IndicatorConfig(**i) for i in indicators])
        db.commit()
        n_cat = db.query(IndicatorCategory).count()
        n_ind = db.query(IndicatorConfig).count()
    print(f"入库完成: indicator_category={n_cat}, indicator_config={n_ind}")


if __name__ == "__main__":
    main()
