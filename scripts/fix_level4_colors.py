# -*- coding: utf-8 -*-
"""修复 xlsx：恢复层级分色样式（从 .bak 恢复原有行颜色），
第4级「具体营业类型」行使用专属浅紫色，且保持树形归位顺序。"""
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill
from collections import OrderedDict

BAK = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系.bak.xlsx"
OUT = r"e:\Project\Web\race\fore-end\docs\农业及相关产业动态指标搜集体系_fixed.xlsx"

# 第4级专属颜色（浅紫），与 黄/蓝/绿/橙 区分
L4_FILL = PatternFill(patternType="solid", fgColor="D9E1F2")


def small_code(cat: str) -> str:
    return str(cat).split("_")[0].split(" ")[0].strip()


def reorder(original, level4):
    last_pos = {}
    for i, r in enumerate(original):
        if r[0].value == "小类":
            last_pos[small_code(r[1].value)] = i
    groups = OrderedDict()
    for r in level4:
        groups.setdefault(small_code(r[1].value), []).append(r)
    insert = {}
    for code, grp in groups.items():
        pos = last_pos.get(code)
        if pos is not None:
            insert.setdefault(pos, []).extend(grp)
        else:
            print(f"[警告] 孤儿: {code}, {len(grp)} 条")
    out = []
    for i, r in enumerate(original):
        out.append(r)
        if i in insert:
            out.extend(insert[i])
    return out


wb = openpyxl.load_workbook(BAK)
ws = wb.active
print(f"读取备份: 数据行 {ws.max_row - 1} 行 x {ws.max_column} 列")

rows = [list(r) for r in ws.iter_rows(min_row=2)]
original = [r for r in rows if r[0].value != "具体营业类型"]
level4 = [r for r in rows if r[0].value == "具体营业类型"]
print(f"原有 {len(original)}, 第4级 {len(level4)}")

new_rows = reorder(original, level4)
print(f"重组后 {len(new_rows)} (应为 {len(original) + len(level4)})")

# 重建：清空数据区，按新顺序写回
ws.delete_rows(2, ws.max_row)
row_idx = 2
l4_written = 0
for r in new_rows:
    is_l4 = r[0].value == "具体营业类型"
    for c, src in enumerate(r, start=1):
        cell = ws.cell(row=row_idx, column=c, value=src.value)
        if is_l4:
            # 第4级：专属浅紫填充（其余用默认样式）
            cell.fill = L4_FILL
        else:
            if src.has_style:
                cell._style = copy(src._style)
    if is_l4:
        l4_written += 1
    row_idx += 1

wb.save(OUT)
print(f"已保存 {OUT}，第4级行 {l4_written}")
